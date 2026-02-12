#!/usr/bin/env python3
"""
Generate the Criticality Analysis Request Form (xlsx) with embedded geometry images.
Each problem type has its own table with specific columns, defaults table, and geometry images.
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


BASE_PATH = Path(__file__).parent.parent

# Problem definitions with columns, defaults, and image paths
PROBLEM_DEFINITIONS = [
    {
        "name": "Single Cylinder",
        "template": "cylinder",
        "description": "Cold traps, RUTS traps, pump cavities",
        "columns": ["#", "Description", "radius_cm", "height_cm", "enrichment", "Notes"],
        "col_widths": [5, 25, 12, 12, 12, 30],
        "defaults": [
            ("Wall material", "Aluminum"),
            ("Wall thickness", "0.32 cm (1/8\")"),
            ("Reflector", "Water, 30 cm"),
            ("UF6 density", "5.09 g/cc"),
        ],
        "num_rows": 4,
        "images": (
            BASE_PATH / "experiments/smoke_test/_validation/geometry.png",
            BASE_PATH / "experiments/smoke_test/_validation/voxel_3d.png",
        ),
    },
    {
        "name": "Rectangular Box",
        "template": "rectangular_box",
        "description": "Chemical traps, HEPA filters, rectangular GEVS",
        "columns": ["#", "Description", "length_cm", "width_cm", "height_cm", "enrichment", "Notes"],
        "col_widths": [5, 25, 12, 12, 12, 12, 30],
        "defaults": [
            ("Wall material", "Steel"),
            ("Wall thickness", "0.32 cm (1/8\")"),
            ("Reflector", "Water, 30 cm"),
            ("UF6 density", "5.09 g/cc"),
        ],
        "num_rows": 4,
        "images": (
            BASE_PATH / "experiments/rectangular_boxes/_validation/geometry.png",
            BASE_PATH / "experiments/rectangular_boxes/_validation/voxel_3d.png",
        ),
    },
    {
        "name": "Process Pipe",
        "template": "process_pipe",
        "description": "Cascade lines, pigtails - single pipe",
        "columns": ["#", "Description", "pipe_size", "length_cm", "enrichment", "Notes"],
        "col_widths": [5, 25, 12, 12, 12, 30],
        "hints": "pipe_size: 1, 1-1/2, 2, 3, 4, 6, 8 (NPS Schedule 10)",
        "defaults": [
            ("Wall material", "SS304"),
            ("Wall thickness", "Per ASME B36.10M"),
            ("Reflector", "Water, 30 cm"),
            ("UF6 density", "5.09 g/cc"),
        ],
        "num_rows": 4,
        "images": (
            BASE_PATH / "experiments/cascade_lines/_validation/geometry.png",
            BASE_PATH / "experiments/cascade_lines/_validation/voxel_3d.png",
        ),
    },
    {
        "name": "Parallel Pipes",
        "template": "parallel_pipes",
        "description": "Piping runs, pipe rack spacing studies",
        "columns": ["#", "Description", "num_pipes", "pipe_size", "length_cm", "gap_cm", "enrichment", "Notes"],
        "col_widths": [5, 25, 12, 12, 12, 12, 12, 25],
        "hints": "num_pipes: 1, 2, or 3 | gap_cm: edge-to-edge spacing",
        "defaults": [
            ("Wall material", "SS304"),
            ("Wall thickness", "Per ASME B36.10M"),
            ("Reflector", "Water, 30 cm"),
            ("UF6 density", "5.09 g/cc"),
        ],
        "num_rows": 4,
        "images": (
            BASE_PATH / "experiments/process_pipes/_validation/geometry.png",
            BASE_PATH / "experiments/process_pipes/_validation/voxel_3d.png",
        ),
    },
    {
        "name": "Shipping Cylinder - Single",
        "template": "shipping_cylinder",
        "description": "Single 5B, 30B, 48Y cylinder",
        "columns": ["#", "Description", "cylinder_type", "enrichment", "Notes"],
        "col_widths": [5, 25, 15, 12, 35],
        "hints": "cylinder_type: 5a, 5b, 30b, 48x, 48y",
        "defaults": [
            ("Wall material", "Per ANSI N14.1 (Monel for 5A/5B, carbon steel for larger)"),
            ("Reflector", "Water, 30 cm"),
            ("Fill fraction", "100%"),
            ("UF6 density", "5.09 g/cc"),
        ],
        "num_rows": 4,
        "images": (
            BASE_PATH / "experiments/benchmarks/uf6_30b/_validation/geometry.png",
            BASE_PATH / "experiments/benchmarks/uf6_30b/_validation/voxel_3d.png",
        ),
    },
    {
        "name": "Cylinder Array - 2D",
        "template": "cylinder_array_2d",
        "description": "Custom traps/vessels arranged in rows x cols",
        "columns": ["#", "Description", "rows", "cols", "radius_cm", "height_cm", "gap_cm", "enrichment", "Notes"],
        "col_widths": [5, 25, 8, 8, 12, 12, 10, 12, 25],
        "hints": "gap_cm: edge-to-edge spacing between cylinder walls",
        "defaults": [
            ("Wall material", "Steel"),
            ("Wall thickness", "0.6 cm"),
            ("Environment", "Air, 30 cm boundary"),
            ("UF6 density", "5.09 g/cc"),
        ],
        "num_rows": 4,
        "images": (
            BASE_PATH / "experiments/cylinder_arrays/_validation/geometry.png",
            BASE_PATH / "experiments/cylinder_arrays/_validation/voxel_3d.png",
        ),
    },
    {
        "name": "Shipping Cylinder Array - 3D",
        "template": "shipping_cylinder_array",
        "description": "Stacked shipping cylinders in warehouse storage",
        "columns": ["#", "Description", "cylinder_type", "orientation", "configuration", "gap_xy_cm", "gap_z_cm", "environment", "enrichment", "Notes"],
        "col_widths": [5, 20, 14, 12, 18, 12, 10, 12, 12, 20],
        "hints": "orientation: vertical or horizontal | configuration: '2x3x2' (RxCxL) for vertical, '3,2,1' (pattern) for horizontal | environment: air or water",
        "defaults": [
            ("Wall material", "Per ANSI N14.1"),
            ("Floor", "Concrete, 30 cm"),
            ("Environment boundary", "30 cm"),
            ("UF6 density", "5.09 g/cc"),
        ],
        "num_rows": 4,
        "images": (
            BASE_PATH / "experiments/stacked_cylinders/_validation/geometry.png",
            BASE_PATH / "experiments/stacked_cylinders/_validation/voxel_3d.png",
        ),
    },
]


def create_form():
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, size=14)
    section_font = Font(bold=True, size=12)
    bold_font = Font(bold=True)
    wrap_alignment = Alignment(wrap_text=True, vertical='top')
    center_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font_white = Font(bold=True, color='FFFFFF', size=10)
    section_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    section_font_white = Font(bold=True, color='FFFFFF', size=12)
    light_blue_fill = PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid')
    defaults_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
    hint_font = Font(italic=True, size=9, color='666666')

    # =========================================================================
    # TAB 1: Instructions
    # =========================================================================
    ws1 = wb.active
    ws1.title = "Instructions"

    ws1['A1'] = "CRITICALITY ANALYSIS REQUEST FORM"
    ws1['A1'].font = Font(bold=True, size=18)
    ws1.merge_cells('A1:E1')

    ws1['A3'] = "Purpose"
    ws1['A3'].font = header_font
    ws1['A4'] = ("Request criticality safety analyses for HALEU equipment. "
                 "Fill out the Request Form tab - we handle the simulation details.")
    ws1['A4'].alignment = wrap_alignment
    ws1.merge_cells('A4:E4')

    ws1['A6'] = "How to Use"
    ws1['A6'].font = header_font
    ws1['A7'] = "1. Go to the 'Request Form' tab"
    ws1['A8'] = "2. Find the section matching your geometry type (each has images to help identify)"
    ws1['A9'] = "3. Fill in dimensions and enrichment in the input table"
    ws1['A10'] = "4. Check the defaults table - add notes if you need different values"

    ws1['A12'] = "Geometry Types Available"
    ws1['A12'].font = header_font

    for i, prob in enumerate(PROBLEM_DEFINITIONS, start=13):
        ws1[f'A{i}'] = f"- {prob['name']}: {prob['description']}"

    ws1.column_dimensions['A'].width = 70
    ws1.column_dimensions['B'].width = 40

    # =========================================================================
    # TAB 2: Request Form (Multiple tables with images)
    # =========================================================================
    ws2 = wb.create_sheet("Request Form")

    ws2['A1'] = "CRITICALITY ANALYSIS REQUESTS"
    ws2['A1'].font = Font(bold=True, size=18)
    ws2.merge_cells('A1:L1')

    ws2['A3'] = "Fill in the appropriate section(s) below. Each section shows geometry images and has its own defaults."
    ws2['A3'].font = Font(italic=True)
    ws2.merge_cells('A3:L3')

    row = 5
    IMAGE_HEIGHT = 180  # pixels
    IMAGE_COL_START = 9  # Column I for images (after the input table)

    for prob in PROBLEM_DEFINITIONS:
        # Section header - spans full width
        num_cols = len(prob["columns"])

        ws2[f'A{row}'] = f"{prob['name'].upper()}"
        ws2[f'A{row}'].font = section_font_white
        ws2[f'A{row}'].fill = section_fill
        for col in range(1, 13):  # Span across all columns
            ws2[f'{get_column_letter(col)}{row}'].fill = section_fill
        ws2.merge_cells(f'A{row}:L{row}')
        ws2.row_dimensions[row].height = 22
        row += 1

        # Description
        ws2[f'A{row}'] = prob["description"]
        ws2[f'A{row}'].font = Font(italic=True, size=10)
        ws2.merge_cells(f'A{row}:H{row}')
        row += 1

        # Hints (if any)
        if "hints" in prob:
            ws2[f'A{row}'] = prob["hints"]
            ws2[f'A{row}'].font = hint_font
            ws2.merge_cells(f'A{row}:H{row}')
            row += 1

        # Track where images should go
        image_start_row = row

        # Column headers for input table
        for i, col_name in enumerate(prob["columns"], start=1):
            cell = ws2[f'{get_column_letter(i)}{row}']
            cell.value = col_name
            cell.font = header_font_white
            cell.fill = header_fill
            cell.alignment = center_alignment
            cell.border = thin_border
        ws2.row_dimensions[row].height = 20
        row += 1

        # Data rows for input
        for r in range(prob["num_rows"]):
            for i in range(1, num_cols + 1):
                cell = ws2[f'{get_column_letter(i)}{row}']
                cell.border = thin_border
                if i == 1:  # Row number column
                    cell.value = r + 1
                    cell.alignment = center_alignment
            row += 1

        # Spacer
        row += 1

        # Defaults table header
        ws2[f'A{row}'] = "Defaults Applied:"
        ws2[f'A{row}'].font = bold_font
        row += 1

        # Defaults table
        for param, value in prob["defaults"]:
            ws2[f'A{row}'] = param
            ws2[f'A{row}'].fill = defaults_fill
            ws2[f'A{row}'].border = thin_border
            ws2[f'B{row}'] = value
            ws2[f'B{row}'].fill = defaults_fill
            ws2[f'B{row}'].border = thin_border
            ws2.merge_cells(f'B{row}:D{row}')
            row += 1

        # Add geometry images to the right of the input table
        img_2d_path, img_3d_path = prob.get("images", (None, None))

        if img_2d_path and img_2d_path.exists():
            try:
                img_2d = Image(str(img_2d_path))
                # Scale image
                scale = IMAGE_HEIGHT / img_2d.height
                img_2d.width = int(img_2d.width * scale)
                img_2d.height = IMAGE_HEIGHT
                ws2.add_image(img_2d, f'I{image_start_row}')
            except Exception as e:
                print(f"  Warning: Could not add 2D image for {prob['name']}: {e}")

        if img_3d_path and img_3d_path.exists():
            try:
                img_3d = Image(str(img_3d_path))
                # Scale image
                scale = IMAGE_HEIGHT / img_3d.height
                img_3d.width = int(img_3d.width * scale)
                img_3d.height = IMAGE_HEIGHT
                ws2.add_image(img_3d, f'K{image_start_row}')
            except Exception as e:
                print(f"  Warning: Could not add 3D image for {prob['name']}: {e}")

        # Ensure enough rows for images (roughly 12 rows per 180px image)
        rows_for_images = IMAGE_HEIGHT // 15
        if row < image_start_row + rows_for_images:
            row = image_start_row + rows_for_images

        # Spacer before next section
        row += 3

    # Set column widths
    col_widths = {
        'A': 5, 'B': 22, 'C': 14, 'D': 12, 'E': 12, 'F': 14,
        'G': 12, 'H': 12, 'I': 25, 'J': 5, 'K': 25, 'L': 5
    }
    for col, width in col_widths.items():
        ws2.column_dimensions[col].width = width

    # =========================================================================
    # TAB 3: Reference - Cylinder Specs
    # =========================================================================
    ws3 = wb.create_sheet("Reference - Cylinders")

    ws3['A1'] = "STANDARD UF6 CYLINDER SPECIFICATIONS"
    ws3['A1'].font = Font(bold=True, size=18)
    ws3.merge_cells('A1:G1')

    ws3['A3'] = "Per ANSI N14.1 - use these cylinder_type values in your request."
    ws3['A3'].font = Font(italic=True)
    ws3.merge_cells('A3:G3')

    cyl_headers = ["Cylinder", "Outer Dia (cm)", "Wall (cm)", "Int. Height (cm)", "Wall Material", "Max Fill (kg)", "Typical Use"]
    for i, h in enumerate(cyl_headers, start=1):
        cell = ws3[f'{get_column_letter(i)}5']
        cell.value = h
        cell.font = bold_font
        cell.fill = light_blue_fill
        cell.border = thin_border

    cyl_data = [
        ("1S", "12.70", "0.16", "22.9", "Steel", "2.2", "Samples"),
        ("2S", "12.70", "0.16", "58.4", "Steel", "5.0", "Samples"),
        ("5A", "12.70", "0.79", "94.0", "Monel", "25.0", "HALEU transport"),
        ("5B", "12.70", "0.79", "94.0", "Monel", "25.0", "HALEU transport"),
        ("30B", "76.20", "1.27", "206.0", "Steel", "2,277", "LEU transport/storage"),
        ("48X", "122.24", "1.59", "302.0", "Steel", "12,501", "Tails storage"),
        ("48Y", "122.24", "1.59", "302.0", "Steel", "12,501", "Feed/product storage"),
        ("48G", "122.24", "1.59", "302.0", "Steel", "12,501", "Heels cylinder"),
    ]
    for i, row_data in enumerate(cyl_data, start=6):
        for j, val in enumerate(row_data, start=1):
            cell = ws3[f'{get_column_letter(j)}{i}']
            cell.value = val
            cell.border = thin_border
            # Highlight common cylinders
            if row_data[0] in ["5A", "5B", "30B", "48Y"]:
                cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    ws3['A15'] = "Note: Highlighted cylinders are most commonly used. Use lowercase in requests (e.g., '48y', '30b')."
    ws3['A15'].font = Font(italic=True, size=10)
    ws3.merge_cells('A15:G15')

    for col, width in {'A': 10, 'B': 14, 'C': 10, 'D': 14, 'E': 14, 'F': 14, 'G': 20}.items():
        ws3.column_dimensions[col].width = width

    # =========================================================================
    # TAB 4: Reference - Pipe Specs
    # =========================================================================
    ws4 = wb.create_sheet("Reference - Pipes")

    ws4['A1'] = "STANDARD PIPE SIZES (Schedule 10/10S)"
    ws4['A1'].font = Font(bold=True, size=18)
    ws4.merge_cells('A1:F1')

    ws4['A3'] = "Per ASME B36.10M / B36.19M - use these pipe_size values in your request."
    ws4['A3'].font = Font(italic=True)
    ws4.merge_cells('A3:F3')

    # Pigtails section
    ws4['A5'] = "Pigtails (Small Diameter)"
    ws4['A5'].font = header_font

    pipe_headers = ["NPS", "OD (cm)", "ID (cm)", "Wall (cm)", "Category", "Typical Use"]
    for i, h in enumerate(pipe_headers, start=1):
        cell = ws4[f'{get_column_letter(i)}6']
        cell.value = h
        cell.font = bold_font
        cell.fill = light_blue_fill
        cell.border = thin_border

    pigtail_data = [
        ("1/8", "1.029", "0.780", "0.124", "Pigtail", "Small connections"),
        ("1/4", "1.372", "1.041", "0.165", "Pigtail", "Small connections"),
        ("3/8", "1.715", "1.384", "0.165", "Pigtail", "Small connections"),
    ]
    for i, row_data in enumerate(pigtail_data, start=7):
        for j, val in enumerate(row_data, start=1):
            cell = ws4[f'{get_column_letter(j)}{i}']
            cell.value = val
            cell.border = thin_border

    # Cascade lines section
    ws4['A11'] = "Cascade Lines (Process Piping)"
    ws4['A11'].font = header_font

    for i, h in enumerate(pipe_headers, start=1):
        cell = ws4[f'{get_column_letter(i)}12']
        cell.value = h
        cell.font = bold_font
        cell.fill = light_blue_fill
        cell.border = thin_border

    cascade_data = [
        ("1", "3.340", "2.786", "0.277", "Cascade", "Process lines"),
        ("1-1/4", "4.216", "3.663", "0.277", "Cascade", "Process lines"),
        ("1-1/2", "4.826", "4.272", "0.277", "Cascade", "Process lines"),
        ("2", "6.032", "5.479", "0.277", "Cascade", "Process lines"),
        ("2-1/2", "7.303", "6.693", "0.305", "Cascade", "Process lines"),
        ("3", "8.890", "8.280", "0.305", "Cascade", "Process lines"),
        ("3-1/2", "10.160", "9.550", "0.305", "Cascade", "Process lines"),
        ("4", "11.430", "10.820", "0.305", "Cascade", "Process lines"),
        ("5", "14.130", "13.449", "0.340", "Cascade", "Headers"),
        ("6", "16.828", "16.147", "0.340", "Cascade", "Headers"),
        ("8", "21.908", "21.156", "0.376", "Cascade", "Large headers"),
    ]
    for i, row_data in enumerate(cascade_data, start=13):
        for j, val in enumerate(row_data, start=1):
            cell = ws4[f'{get_column_letter(j)}{i}']
            cell.value = val
            cell.border = thin_border
            # Highlight common sizes
            if row_data[0] in ["2", "3", "4"]:
                cell.fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')

    ws4['A25'] = "Note: Highlighted sizes are most common. Use quotes for fractions (e.g., '1-1/4', '2')."
    ws4['A25'].font = Font(italic=True, size=10)
    ws4.merge_cells('A25:F25')

    for col, width in {'A': 8, 'B': 10, 'C': 10, 'D': 10, 'E': 10, 'F': 18}.items():
        ws4.column_dimensions[col].width = width

    # =========================================================================
    # TAB 5: Reference - Materials
    # =========================================================================
    ws5 = wb.create_sheet("Reference - Materials")

    ws5['A1'] = "AVAILABLE MATERIALS"
    ws5['A1'].font = Font(bold=True, size=18)
    ws5.merge_cells('A1:D1')

    # Wall materials
    ws5['A3'] = "Wall Materials"
    ws5['A3'].font = header_font

    wall_headers = ["Material", "Keyword", "Density (g/cc)", "Typical Use"]
    for i, h in enumerate(wall_headers, start=1):
        cell = ws5[f'{get_column_letter(i)}4']
        cell.value = h
        cell.font = bold_font
        cell.fill = light_blue_fill
        cell.border = thin_border

    wall_data = [
        ("Carbon Steel", "steel", "7.82", "Process piping, vessels, 30B/48Y cylinders"),
        ("Stainless Steel 304", "ss304", "8.03", "Corrosion-resistant pipes, process equipment"),
        ("Aluminum", "aluminum", "2.70", "Lightweight containers"),
        ("Monel 400", "monel", "8.80", "5A/5B cylinders (HALEU)"),
    ]
    for i, row_data in enumerate(wall_data, start=5):
        for j, val in enumerate(row_data, start=1):
            cell = ws5[f'{get_column_letter(j)}{i}']
            cell.value = val
            cell.border = thin_border

    # Reflector materials
    ws5['A11'] = "Reflector / Environment Materials"
    ws5['A11'].font = header_font

    refl_headers = ["Material", "Keyword", "Density (g/cc)", "Notes"]
    for i, h in enumerate(refl_headers, start=1):
        cell = ws5[f'{get_column_letter(i)}12']
        cell.value = h
        cell.font = bold_font
        cell.fill = light_blue_fill
        cell.border = thin_border

    refl_data = [
        ("Water", "water", "1.00", "Most reactive - conservative default for reflection"),
        ("Concrete", "concrete", "2.30", "Building floors, walls - use for stacked cylinder floor"),
        ("Air", "air", "0.001", "Normal environment - minimal moderation"),
        ("None", "none", "-", "Bare/unreflected (vacuum boundary condition)"),
    ]
    for i, row_data in enumerate(refl_data, start=13):
        for j, val in enumerate(row_data, start=1):
            cell = ws5[f'{get_column_letter(j)}{i}']
            cell.value = val
            cell.border = thin_border

    ws5['A19'] = "Default: Water reflection is used unless otherwise specified (conservative assumption)."
    ws5['A19'].font = Font(italic=True, size=10)
    ws5.merge_cells('A19:D19')

    ws5['A20'] = "For stacked cylinder studies, use 'air' environment for normal storage or 'water' for flooding scenarios."
    ws5['A20'].font = Font(italic=True, size=10)
    ws5.merge_cells('A20:D20')

    for col, width in {'A': 18, 'B': 12, 'C': 14, 'D': 45}.items():
        ws5.column_dimensions[col].width = width

    # =========================================================================
    # TAB 6: Reference - Stacking Patterns
    # =========================================================================
    ws6 = wb.create_sheet("Reference - Stacking")

    ws6['A1'] = "CYLINDER STACKING PATTERNS"
    ws6['A1'].font = Font(bold=True, size=18)
    ws6.merge_cells('A1:D1')

    ws6['A3'] = "For shipping_cylinder_stacked template - specify stacking_pattern as comma-separated values."
    ws6['A3'].font = Font(italic=True)
    ws6.merge_cells('A3:D3')

    ws6['A5'] = "Common Patterns"
    ws6['A5'].font = header_font

    stack_headers = ["Pattern", "Total Cylinders", "Description", "Visual"]
    for i, h in enumerate(stack_headers, start=1):
        cell = ws6[f'{get_column_letter(i)}6']
        cell.value = h
        cell.font = bold_font
        cell.fill = light_blue_fill
        cell.border = thin_border

    stack_data = [
        ("3,2,1", "6", "Pyramid - 3 bottom, 2 middle, 1 top", "  o\n oo\nooo"),
        ("2,1", "3", "Simple pyramid - 2 bottom, 1 top", " o\noo"),
        ("3,3,3", "9", "Rectangular 3-high stack", "ooo\nooo\nooo"),
        ("2,2,2", "6", "Rectangular 2-wide, 3-high", "oo\noo\noo"),
        ("3,3", "6", "Rectangular 3-wide, 2-high", "ooo\nooo"),
        ("1", "1", "Single cylinder", "o"),
        ("2", "2", "Two side-by-side", "oo"),
        ("3", "3", "Three side-by-side", "ooo"),
    ]
    for i, row_data in enumerate(stack_data, start=7):
        for j, val in enumerate(row_data, start=1):
            cell = ws6[f'{get_column_letter(j)}{i}']
            cell.value = val
            cell.border = thin_border
            if j == 4:  # Visual column
                cell.alignment = Alignment(wrap_text=True, vertical='top')
                cell.font = Font(name='Courier New', size=9)

    ws6['A17'] = "Note: Bottom layer is listed first. Cylinders are horizontal (lying on their side)."
    ws6['A17'].font = Font(italic=True, size=10)
    ws6.merge_cells('A17:D17')

    ws6['A18'] = "Gap between cylinders controlled by gap_y_cm (same layer) and gap_z_cm (between layers)."
    ws6['A18'].font = Font(italic=True, size=10)
    ws6.merge_cells('A18:D18')

    for col, width in {'A': 12, 'B': 16, 'C': 35, 'D': 12}.items():
        ws6.column_dimensions[col].width = width

    # Save
    output_path = BASE_PATH / "docs" / "Criticality_Request_Form_v3.xlsx"
    wb.save(output_path)
    print(f"Created: {output_path}")

    # Print image status
    print("\nImage embedding status:")
    for prob in PROBLEM_DEFINITIONS:
        img_2d, img_3d = prob.get("images", (None, None))
        status_2d = "OK" if img_2d and img_2d.exists() else "MISSING"
        status_3d = "OK" if img_3d and img_3d.exists() else "MISSING"
        print(f"  {prob['name']}: 2D={status_2d}, 3D={status_3d}")

    return output_path


if __name__ == "__main__":
    create_form()
