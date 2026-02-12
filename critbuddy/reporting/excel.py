"""
Excel report generator for criticality experiment results.

Generates XLSX workbooks with:
- Multiple data sheets for different conditions
- Embedded heatmap and line charts
- Conditional formatting for safety status
- Summary tables with key findings
"""

from pathlib import Path
from typing import Dict, List, Optional, Union
import pandas as pd

try:
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side,
        NamedStyle
    )
    from openpyxl.chart import (
        BarChart, LineChart, Reference, Series
    )
    from openpyxl.chart.label import DataLabelList
    from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False


# Style definitions
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SAFE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MARGINAL_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
CRITICAL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def generate_lookup_xlsx(
    run_dirs: Dict[str, Union[str, Path]],
    output_path: Union[str, Path],
    experiment_name: str = "Cascade Array",
) -> Path:
    """
    Generate Excel look-up table workbook from experiment results.

    Args:
        run_dirs: Dict mapping condition names to run directories
                  e.g., {"optimal": "/path/to/0.5_run", "flooded": "/path/to/1.0_run"}
        output_path: Path for output XLSX file
        experiment_name: Name to display in report headers

    Returns:
        Path to generated XLSX file
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError(
            "openpyxl is required for Excel generation. "
            "Install with: pip install openpyxl"
        )

    output_path = Path(output_path)

    # Load all results
    results = {}
    for condition, run_dir in run_dirs.items():
        run_dir = Path(run_dir)
        csv_path = run_dir / "results.csv"
        if csv_path.exists():
            results[condition] = pd.read_csv(csv_path)
        else:
            raise FileNotFoundError(f"Results not found: {csv_path}")

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets
    _create_summary_sheet(wb, results, experiment_name)

    for condition, df in results.items():
        _create_results_sheet(wb, df, condition.replace("_", " ").title())

    if len(results) >= 2:
        _create_comparison_sheet(wb, results)

    _create_geometry_sheet(wb, results)

    # Save workbook
    wb.save(output_path)

    return output_path


def _create_summary_sheet(wb: Workbook, results: Dict[str, pd.DataFrame], name: str):
    """Create summary sheet with key findings."""
    ws = wb.create_sheet("Summary")

    # Title
    ws["A1"] = f"Criticality Analysis: {name}"
    ws["A1"].font = Font(bold=True, size=16)
    ws.merge_cells("A1:F1")

    row = 3

    # Key findings for each condition
    for condition, df in results.items():
        ws.cell(row=row, column=1, value=f"Condition: {condition.replace('_', ' ').title()}")
        ws.cell(row=row, column=1).font = Font(bold=True, size=12)
        row += 1

        # Statistics
        n_safe = len(df[df["status"] == "SAFE"])
        n_marginal = len(df[df["status"] == "MARGINAL"])
        n_critical = len(df[df["status"] == "CRITICAL"])
        n_total = len(df)

        stats = [
            ("Total Cases", n_total),
            ("SAFE", f"{n_safe} ({100*n_safe/n_total:.0f}%)"),
            ("MARGINAL", f"{n_marginal} ({100*n_marginal/n_total:.0f}%)"),
            ("CRITICAL", f"{n_critical} ({100*n_critical/n_total:.0f}%)"),
            ("k-eff Range", f"{df['keff'].min():.4f} - {df['keff'].max():.4f}"),
        ]

        for label, value in stats:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1

        row += 1

    # Minimum safe gaps table
    row += 1
    ws.cell(row=row, column=1, value="Minimum Safe Gap by Enrichment")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12)
    row += 1

    # Headers
    headers = ["Enrichment (%)", "Min Safe Gap (cm)", "k-eff at Min Gap", "Status"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Find minimum safe gap for each enrichment (use first condition with critical cases)
    primary_df = list(results.values())[0]
    if "enrichment" in primary_df.columns and "gap_xy_cm" in primary_df.columns:
        for enr in sorted(primary_df["enrichment"].unique()):
            enr_data = primary_df[primary_df["enrichment"] == enr]
            safe_data = enr_data[enr_data["status"] == "SAFE"]

            if len(safe_data) > 0:
                min_gap_row = safe_data.loc[safe_data["gap_xy_cm"].idxmin()]
                min_gap = min_gap_row["gap_xy_cm"]
                keff = min_gap_row["keff"]
                status = "SAFE"
            else:
                # All critical
                min_gap_row = enr_data.loc[enr_data["keff"].idxmin()]
                min_gap = f"> {enr_data['gap_xy_cm'].max()}"
                keff = min_gap_row["keff"]
                status = "CRITICAL"

            ws.cell(row=row, column=1, value=enr)
            ws.cell(row=row, column=2, value=min_gap)
            ws.cell(row=row, column=3, value=f"{keff:.4f}")
            status_cell = ws.cell(row=row, column=4, value=status)
            _apply_status_fill(status_cell, status)
            row += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 12


def _create_results_sheet(wb: Workbook, df: pd.DataFrame, sheet_name: str):
    """Create results sheet with data table and charts."""
    ws = wb.create_sheet(sheet_name)

    # Identify swept and fixed params
    standard_cols = {"case", "solver", "keff", "std", "keff_2sigma", "status", "execution_time"}
    param_cols = [c for c in df.columns if c not in standard_cols]

    swept_params = {}
    for col in param_cols:
        unique_vals = df[col].unique()
        if len(unique_vals) > 1:
            swept_params[col] = sorted(unique_vals.tolist())

    # Write header
    ws["A1"] = f"Results: {sheet_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")

    # Write pivot table if 2D sweep
    row = 3
    if len(swept_params) >= 2:
        param_names = list(swept_params.keys())
        row_param = param_names[0]  # enrichment
        col_param = param_names[1]  # gap

        # Create pivot
        pivot = df.pivot_table(
            values="keff",
            index=row_param,
            columns=col_param,
            aggfunc="first"
        )

        # Write pivot table header
        ws.cell(row=row, column=1, value=f"k-eff: {row_param} vs {col_param}")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        # Column headers
        ws.cell(row=row, column=1, value=row_param)
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=1).font = HEADER_FONT
        for col_idx, col_val in enumerate(pivot.columns, start=2):
            cell = ws.cell(row=row, column=col_idx, value=col_val)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Data rows
        pivot_start_row = row
        for row_val in pivot.index:
            ws.cell(row=row, column=1, value=row_val)
            ws.cell(row=row, column=1).fill = HEADER_FILL
            ws.cell(row=row, column=1).font = HEADER_FONT
            for col_idx, col_val in enumerate(pivot.columns, start=2):
                keff_val = pivot.loc[row_val, col_val]
                cell = ws.cell(row=row, column=col_idx, value=round(keff_val, 4))
                cell.number_format = "0.0000"
                cell.alignment = Alignment(horizontal="center")

                # Color based on value
                if keff_val >= 1.0:
                    cell.fill = CRITICAL_FILL
                elif keff_val >= 0.95:
                    cell.fill = MARGINAL_FILL
                else:
                    cell.fill = SAFE_FILL
            row += 1
        pivot_end_row = row - 1

        # Add color scale rule
        color_range = f"B{pivot_start_row}:{get_column_letter(1 + len(pivot.columns))}{pivot_end_row}"

        row += 2

        # Status pivot table
        status_pivot = df.pivot_table(
            values="status",
            index=row_param,
            columns=col_param,
            aggfunc="first"
        )

        ws.cell(row=row, column=1, value=f"Status: {row_param} vs {col_param}")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        # Column headers
        ws.cell(row=row, column=1, value=row_param)
        ws.cell(row=row, column=1).fill = HEADER_FILL
        ws.cell(row=row, column=1).font = HEADER_FONT
        for col_idx, col_val in enumerate(status_pivot.columns, start=2):
            cell = ws.cell(row=row, column=col_idx, value=col_val)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Data rows
        for row_val in status_pivot.index:
            ws.cell(row=row, column=1, value=row_val)
            ws.cell(row=row, column=1).fill = HEADER_FILL
            ws.cell(row=row, column=1).font = HEADER_FONT
            for col_idx, col_val in enumerate(status_pivot.columns, start=2):
                status = status_pivot.loc[row_val, col_val]
                cell = ws.cell(row=row, column=col_idx, value=status)
                cell.alignment = Alignment(horizontal="center")
                _apply_status_fill(cell, status)
            row += 1

    row += 2

    # Full data table
    ws.cell(row=row, column=1, value="Full Results")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    # Select columns to display
    display_cols = ["case", "keff", "std", "keff_2sigma", "status"] + list(swept_params.keys())
    display_df = df[display_cols].copy()

    # Headers
    for col_idx, col_name in enumerate(display_df.columns, start=1):
        cell = ws.cell(row=row, column=col_idx, value=col_name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    row += 1

    # Data
    for _, data_row in display_df.iterrows():
        for col_idx, col_name in enumerate(display_df.columns, start=1):
            val = data_row[col_name]
            cell = ws.cell(row=row, column=col_idx, value=val)
            if col_name in ["keff", "std", "keff_2sigma"]:
                cell.number_format = "0.00000"
            if col_name == "status":
                _apply_status_fill(cell, val)
            cell.alignment = Alignment(horizontal="center")
        row += 1

    # Adjust column widths
    for col_idx in range(1, len(display_df.columns) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 14


def _create_comparison_sheet(wb: Workbook, results: Dict[str, pd.DataFrame]):
    """Create comparison sheet showing difference between conditions."""
    ws = wb.create_sheet("Comparison")

    conditions = list(results.keys())
    if len(conditions) < 2:
        return

    cond1, cond2 = conditions[0], conditions[1]
    df1, df2 = results[cond1], results[cond2]

    ws["A1"] = f"Comparison: {cond1.title()} vs {cond2.title()}"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:G1")

    row = 3

    # Identify common parameters
    standard_cols = {"case", "solver", "keff", "std", "keff_2sigma", "status", "execution_time"}
    param_cols = [c for c in df1.columns if c not in standard_cols]

    swept_params = {}
    for col in param_cols:
        if col in df1.columns and col in df2.columns:
            unique_vals = df1[col].unique()
            if len(unique_vals) > 1:
                swept_params[col] = sorted(unique_vals.tolist())

    # Create merged comparison dataframe
    if len(swept_params) >= 2:
        merge_cols = list(swept_params.keys())

        merged = df1[merge_cols + ["keff"]].merge(
            df2[merge_cols + ["keff"]],
            on=merge_cols,
            suffixes=(f"_{cond1}", f"_{cond2}")
        )
        merged["delta_k"] = merged[f"keff_{cond1}"] - merged[f"keff_{cond2}"]

        # Headers
        headers = merge_cols + [f"k-eff ({cond1})", f"k-eff ({cond2})", "Δk"]
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col_idx, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
        row += 1

        # Data
        for _, data_row in merged.iterrows():
            col_idx = 1
            for col in merge_cols:
                ws.cell(row=row, column=col_idx, value=data_row[col])
                col_idx += 1

            cell1 = ws.cell(row=row, column=col_idx, value=data_row[f"keff_{cond1}"])
            cell1.number_format = "0.0000"
            col_idx += 1

            cell2 = ws.cell(row=row, column=col_idx, value=data_row[f"keff_{cond2}"])
            cell2.number_format = "0.0000"
            col_idx += 1

            delta_cell = ws.cell(row=row, column=col_idx, value=data_row["delta_k"])
            delta_cell.number_format = "+0.000;-0.000;0.000"
            # Color delta: positive = red (cond1 higher), negative = green
            if data_row["delta_k"] > 0.1:
                delta_cell.fill = CRITICAL_FILL
            elif data_row["delta_k"] > 0:
                delta_cell.fill = MARGINAL_FILL
            else:
                delta_cell.fill = SAFE_FILL

            row += 1

        # Summary statistics
        row += 2
        ws.cell(row=row, column=1, value="Summary Statistics")
        ws.cell(row=row, column=1).font = Font(bold=True)
        row += 1

        stats = [
            ("Mean Δk", merged["delta_k"].mean()),
            ("Min Δk", merged["delta_k"].min()),
            ("Max Δk", merged["delta_k"].max()),
        ]
        for label, val in stats:
            ws.cell(row=row, column=1, value=label)
            cell = ws.cell(row=row, column=2, value=val)
            cell.number_format = "+0.0000;-0.0000;0.0000"
            row += 1

    # Adjust column widths
    for col_idx in range(1, 8):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16


def _create_geometry_sheet(wb: Workbook, results: Dict[str, pd.DataFrame]):
    """Create geometry specification sheet."""
    ws = wb.create_sheet("Geometry")

    ws["A1"] = "Array Configuration"
    ws["A1"].font = Font(bold=True, size=14)

    # Get params from first result set
    df = list(results.values())[0]

    row = 3
    # Extract geometry params from first row
    first_row = df.iloc[0]

    geometry_params = [
        ("rows", "Rows (X direction)"),
        ("cols", "Columns (Y direction)"),
        ("layers", "Layers (Z direction)"),
        ("radius_cm", "Cylinder radius (cm)"),
        ("height_cm", "Cylinder height (cm)"),
        ("wall_thickness_cm", "Wall thickness (cm)"),
        ("wall_material", "Wall material"),
        ("water_thickness_cm", "Reflector thickness (cm)"),
    ]

    headers = ["Parameter", "Value"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    row += 1

    for param, label in geometry_params:
        if param in df.columns:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=first_row[param])
            row += 1

    # Calculate derived values
    row += 1
    ws.cell(row=row, column=1, value="Derived Dimensions")
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 1

    if all(p in df.columns for p in ["rows", "cols", "layers"]):
        total_cyl = first_row["rows"] * first_row["cols"] * first_row["layers"]
        ws.cell(row=row, column=1, value="Total cylinders")
        ws.cell(row=row, column=2, value=total_cyl)
        row += 1

    if all(p in df.columns for p in ["radius_cm", "wall_thickness_cm"]):
        outer_r = first_row["radius_cm"] + first_row["wall_thickness_cm"]
        ws.cell(row=row, column=1, value="Outer radius (cm)")
        ws.cell(row=row, column=2, value=round(outer_r, 4))
        row += 1

        inner_d = 2 * first_row["radius_cm"]
        outer_d = 2 * outer_r
        ws.cell(row=row, column=1, value="Inner diameter (cm)")
        ws.cell(row=row, column=2, value=round(inner_d, 2))
        row += 1
        ws.cell(row=row, column=1, value="Outer diameter (cm)")
        ws.cell(row=row, column=2, value=round(outer_d, 2))
        row += 1

    # Adjust column widths
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 18


def _apply_status_fill(cell, status: str):
    """Apply fill color based on safety status."""
    if status == "SAFE":
        cell.fill = SAFE_FILL
    elif status == "MARGINAL":
        cell.fill = MARGINAL_FILL
    elif status == "CRITICAL":
        cell.fill = CRITICAL_FILL


def generate_cascade_array_xlsx(
    optimal_run_dir: Union[str, Path],
    flooded_run_dir: Union[str, Path],
    output_path: Union[str, Path],
) -> Path:
    """
    Convenience function for cascade array experiments.

    Args:
        optimal_run_dir: Run directory for optimal moderation (0.5 g/cc)
        flooded_run_dir: Run directory for flooded condition (1.0 g/cc)
        output_path: Output XLSX path

    Returns:
        Path to generated XLSX
    """
    return generate_lookup_xlsx(
        run_dirs={
            "optimal_moderation": optimal_run_dir,
            "flooded": flooded_run_dir,
        },
        output_path=output_path,
        experiment_name="Cascade Array (10×5×4)",
    )
